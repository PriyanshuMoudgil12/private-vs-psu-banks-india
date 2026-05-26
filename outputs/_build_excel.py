"""Build the executive Excel summary for the project."""
from pathlib import Path
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

HERE = Path(__file__).resolve().parent
DATA = HERE.parent / "data" / "bank_financials.csv"
OUT  = HERE / "Private_vs_PSU_Banks_Executive_Summary.xlsx"

df = pd.read_csv(DATA)

# ------------------------------------------------------------------ styles
FONT = "Arial"
BOLD   = Font(name=FONT, bold=True, size=11)
HEADER = Font(name=FONT, bold=True, size=11, color="FFFFFF")
TITLE  = Font(name=FONT, bold=True, size=16, color="1F2A44")
ITAL   = Font(name=FONT, italic=True, size=10, color="555555")
NORM   = Font(name=FONT, size=11)
NAVY   = PatternFill("solid", start_color="1F2A44")
BLUE   = PatternFill("solid", start_color="DCE6F1")
RED    = PatternFill("solid", start_color="F4CCCC")
GREEN  = PatternFill("solid", start_color="D9EAD3")
YELLOW = PatternFill("solid", start_color="FFF2CC")
GREY   = PatternFill("solid", start_color="F3F3F3")

CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
LEFT   = Alignment(horizontal="left",   vertical="center", wrap_text=True)
THIN   = Side(style="thin", color="BFBFBF")
BORDER = Border(top=THIN, left=THIN, right=THIN, bottom=THIN)

def style_header_row(ws, row, n_cols, fill=NAVY, font=HEADER):
    for c in range(1, n_cols + 1):
        cell = ws.cell(row=row, column=c)
        cell.fill = fill
        cell.font = font
        cell.alignment = CENTER
        cell.border = BORDER

def autosize(ws, min_w=10, max_w=30):
    for col in ws.columns:
        col_letter = get_column_letter(col[0].column)
        width = min_w
        for cell in col:
            if cell.value is not None:
                width = max(width, min(max_w, len(str(cell.value)) + 2))
        ws.column_dimensions[col_letter].width = width

wb = Workbook()

# ===================== SHEET 1 — EXECUTIVE SUMMARY =====================
ws = wb.active
ws.title = "Executive Summary"

ws["A1"] = "Private vs PSU Banks in India — Strategic Analysis (FY20–FY25)"
ws["A1"].font = TITLE
ws.merge_cells("A1:F1")

ws["A2"] = ("6 banks · 6 years · ₹50+ lakh crore of assets. Prepared by Priyanshu Moudgil · "
            "BBA · Business Analyst Portfolio.")
ws["A2"].font = ITAL
ws.merge_cells("A2:F2")

ws["A4"] = "The Big Picture (FY25)"
ws["A4"].font = BOLD
ws["A4"].fill = YELLOW
ws.merge_cells("A4:F4")

headers = ["Metric", "Private avg", "PSU avg", "Gap (pp)", "Direction", "Take-away"]
for i, h in enumerate(headers, 1):
    ws.cell(row=5, column=i, value=h)
style_header_row(ws, 5, len(headers))

# These values come from groupby averages on the FY25 data — wired as formulas
# in the Raw Data sheet via AVERAGEIFS so the summary stays dynamic.
rows = [
    ("ROA (%)",              "=AVERAGEIFS('Raw Data'!I:I,'Raw Data'!B:B,\"Private\",'Raw Data'!C:C,\"FY25\")",
                              "=AVERAGEIFS('Raw Data'!I:I,'Raw Data'!B:B,\"PSU\",'Raw Data'!C:C,\"FY25\")",
                              "=B6-C6", "↑ widening",
                              "Private banks earn ~2× more per ₹ of assets."),
    ("Gross NPA (%)",        "=AVERAGEIFS('Raw Data'!H:H,'Raw Data'!B:B,\"Private\",'Raw Data'!C:C,\"FY25\")",
                              "=AVERAGEIFS('Raw Data'!H:H,'Raw Data'!B:B,\"PSU\",'Raw Data'!C:C,\"FY25\")",
                              "=C7-B7", "↓ narrowing",
                              "PSU NPA clean-up cut excess NPA from 6pp to 1pp."),
    ("NIM (%)",              "=AVERAGEIFS('Raw Data'!J:J,'Raw Data'!B:B,\"Private\",'Raw Data'!C:C,\"FY25\")",
                              "=AVERAGEIFS('Raw Data'!J:J,'Raw Data'!B:B,\"PSU\",'Raw Data'!C:C,\"FY25\")",
                              "=B8-C8", "→ stable",
                              "~90bps structural spread on the lending side."),
    ("CASA (%)",             "=AVERAGEIFS('Raw Data'!K:K,'Raw Data'!B:B,\"Private\",'Raw Data'!C:C,\"FY25\")",
                              "=AVERAGEIFS('Raw Data'!K:K,'Raw Data'!B:B,\"PSU\",'Raw Data'!C:C,\"FY25\")",
                              "=B9-C9", "↓ flipped",
                              "FY25: PSU CASA edges past private for the first time."),
]
for r, row_vals in enumerate(rows, 6):
    for c, v in enumerate(row_vals, 1):
        cell = ws.cell(row=r, column=c, value=v)
        cell.font = NORM
        cell.border = BORDER
        cell.alignment = LEFT if c in (1, 5, 6) else CENTER
        if c in (2, 3, 4):
            cell.number_format = "0.00"

# ----- The 5 findings
ws["A11"] = "5 Key Findings"
ws["A11"].font = BOLD
ws["A11"].fill = YELLOW
ws.merge_cells("A11:F11")

findings = [
    ("#1", "ROA gap held at ~1 percentage point.", "Private ROA doubled (0.97% → 1.98%). PSU ROA grew 7× from a tiny base (0.16% → 1.12%). The gap is structural."),
    ("#2", "The PSU NPA clean-up is real — and almost over.", "PSU gross NPA fell from 9.92% → 2.68%. Excess vs private fell from 6.04pp to 1.23pp. The easy gains are done."),
    ("#3", "PSUs aren't growth-laggards, they're just smaller.", "PNB loan-book CAGR 18.8%, SBI 12.7% — same range as ICICI/Axis. The 'PSUs are dying' narrative is wrong."),
    ("#4", "The CASA advantage just evaporated.", "Private CASA peaked at 47.4% (FY22) → 38.1% (FY25). In FY25 PSUs (39.6%) actually beat private."),
    ("#5", "NIM: the persistent ~90bps spread.", "Private banks earn ~90bps more on every rupee they lend — pricing power & retail mix, not luck."),
]
for r, (num, title, body) in enumerate(findings, 12):
    ws.cell(row=r, column=1, value=num).font = BOLD
    ws.cell(row=r, column=2, value=title).font = BOLD
    ws.cell(row=r, column=3, value=body).font = NORM
    ws.merge_cells(start_row=r, start_column=3, end_row=r, end_column=6)
    for c in range(1, 7):
        ws.cell(row=r, column=c).border = BORDER
        ws.cell(row=r, column=c).alignment = LEFT

# ----- 3 recommendations
ws["A18"] = "3 Strategic Recommendations for the PSU CEO"
ws["A18"].font = BOLD
ws["A18"].fill = GREEN
ws.merge_cells("A18:F18")

recs = [
    ("1", "Stop chasing CASA. Chase NIM.",
     "The CASA war is over and the rate cycle decided it. Shift 5pp of the loan book from corporate to high-yield retail by FY27; stand up a digital-lending JV."),
    ("2", "Defend NPA gains by re-engineering underwriting, not lending less.",
     "IBC won the first NPA war. Win the second with data: bureau scores, GST flows, real-time monitoring. Mandate model-driven decisions for loans under ₹5cr."),
    ("3", "Become the bank for Bharat's middle 60%.",
     "600M middle-income, semi-urban, GST-registered SMBs are under-served. PSUs have the trust and branches — close the digital UX gap. Target ₹5 lakh crore new SMB advances by FY28."),
]
for r, (num, title, body) in enumerate(recs, 19):
    ws.cell(row=r, column=1, value=num).font = BOLD
    ws.cell(row=r, column=2, value=title).font = BOLD
    ws.cell(row=r, column=3, value=body).font = NORM
    ws.merge_cells(start_row=r, start_column=3, end_row=r, end_column=6)
    for c in range(1, 7):
        ws.cell(row=r, column=c).border = BORDER
        ws.cell(row=r, column=c).alignment = LEFT
    ws.row_dimensions[r].height = 36

ws.column_dimensions["A"].width = 6
ws.column_dimensions["B"].width = 38
ws.column_dimensions["C"].width = 22
ws.column_dimensions["D"].width = 12
ws.column_dimensions["E"].width = 14
ws.column_dimensions["F"].width = 36

# ===================== SHEET 2 — RAW DATA =====================
ws = wb.create_sheet("Raw Data")
ws["A1"] = "Bank-year master dataset (FY20–FY25)"
ws["A1"].font = TITLE
ws.merge_cells("A1:K1")
ws["A2"] = ("Absolute values in ₹ crore. Ratios in percent. Source: each bank's annual report + "
            "investor presentations + Equitymaster + verified via May 2026 web search.")
ws["A2"].font = ITAL
ws.merge_cells("A2:K2")

cols = list(df.columns)
for c, name in enumerate(cols, 1):
    ws.cell(row=4, column=c, value=name)
style_header_row(ws, 4, len(cols))
for r, row in enumerate(df.itertuples(index=False), 5):
    for c, val in enumerate(row, 1):
        cell = ws.cell(row=r, column=c, value=val)
        cell.font = NORM
        cell.border = BORDER
        cell.alignment = LEFT if c <= 3 else CENTER
        if c >= 4 and c <= 7:
            cell.number_format = "#,##0"
        if c >= 8:
            cell.number_format = "0.00"
    # row tint by type
    if row.bank_type == "Private":
        for c in range(1, len(cols) + 1):
            ws.cell(row=r, column=c).fill = BLUE
    else:
        for c in range(1, len(cols) + 1):
            ws.cell(row=r, column=c).fill = RED
autosize(ws, min_w=10, max_w=20)

# ===================== SHEET 3 — GROUP AVERAGES =====================
ws = wb.create_sheet("Group Averages")
ws["A1"] = "Year-by-year averages — Private vs PSU"
ws["A1"].font = TITLE
ws.merge_cells("A1:H1")

# 4 mini tables: ROA, GNPA, NIM, CASA — all driven by AVERAGEIFS formulas
years = ["FY20", "FY21", "FY22", "FY23", "FY24", "FY25"]
blocks = [
    ("ROA (%)",       "I", "0.00"),
    ("Gross NPA (%)", "H", "0.00"),
    ("NIM (%)",       "J", "0.00"),
    ("CASA (%)",      "K", "0.00"),
]
start_row = 3
for metric, col_letter, fmt in blocks:
    ws.cell(row=start_row, column=1, value=metric).font = BOLD
    ws.cell(row=start_row, column=1).fill = YELLOW
    ws.merge_cells(start_row=start_row, start_column=1,
                   end_row=start_row, end_column=4)

    ws.cell(row=start_row+1, column=1, value="Year").font = HEADER
    ws.cell(row=start_row+1, column=2, value="Private").font = HEADER
    ws.cell(row=start_row+1, column=3, value="PSU").font = HEADER
    ws.cell(row=start_row+1, column=4, value="Gap (pp)").font = HEADER
    for c in range(1, 5):
        ws.cell(row=start_row+1, column=c).fill = NAVY
        ws.cell(row=start_row+1, column=c).alignment = CENTER
        ws.cell(row=start_row+1, column=c).border = BORDER

    for i, yr in enumerate(years):
        r = start_row + 2 + i
        ws.cell(row=r, column=1, value=yr).alignment = CENTER
        ws.cell(row=r, column=2,
                value=f"=AVERAGEIFS('Raw Data'!{col_letter}:{col_letter},"
                      f"'Raw Data'!B:B,\"Private\",'Raw Data'!C:C,\"{yr}\")")
        ws.cell(row=r, column=3,
                value=f"=AVERAGEIFS('Raw Data'!{col_letter}:{col_letter},"
                      f"'Raw Data'!B:B,\"PSU\",'Raw Data'!C:C,\"{yr}\")")
        ws.cell(row=r, column=4, value=f"=B{r}-C{r}")
        for c in range(1, 5):
            cell = ws.cell(row=r, column=c)
            cell.number_format = fmt if c > 1 else "@"
            cell.border = BORDER
            cell.font = NORM
            if c > 1:
                cell.alignment = CENTER
    start_row += 10

for col in "ABCDE":
    ws.column_dimensions[col].width = 16

# ===================== SHEET 4 — FY25 LEAGUE TABLE =====================
ws = wb.create_sheet("FY25 League Table")
ws["A1"] = "FY25 — All 6 banks ranked best-to-worst"
ws["A1"].font = TITLE
ws.merge_cells("A1:I1")

hdrs = ["Rank", "Bank", "Type", "Total assets (₹cr)", "Net profit (₹cr)",
        "ROA (%)", "Gross NPA (%)", "NIM (%)", "CASA (%)"]
for c, h in enumerate(hdrs, 1):
    ws.cell(row=3, column=c, value=h)
style_header_row(ws, 3, len(hdrs))

# Hardcoded from the SQL Q6 output — final ranking
league = [
    (1, "Axis Bank",       "Private", 1610327, 28055, 1.88, 1.30, 3.98, 41.00),
    (2, "ICICI Bank",      "Private", 1918233, 47227, 2.46, 1.70, 4.40, 38.40),
    (3, "HDFC Bank",       "Private", 4392420, 70792, 1.61, 1.33, 3.90, 34.90),
    (3, "SBI",             "PSU",     6606290, 70901, 1.10, 1.82, 3.09, 39.97),
    (5, "Bank of Baroda",  "PSU",     1710000, 20716, 1.26, 2.26, 3.02, 39.95),
    (6, "PNB",             "PSU",     1650000, 16630, 1.00, 3.95, 2.93, 38.80),
]
for i, row in enumerate(league, 4):
    for c, v in enumerate(row, 1):
        cell = ws.cell(row=i, column=c, value=v)
        cell.font = NORM
        cell.border = BORDER
        cell.alignment = LEFT if c == 2 else CENTER
        if c in (4, 5):
            cell.number_format = "#,##0"
        if c in (6, 7, 8, 9):
            cell.number_format = "0.00"
    fill = BLUE if row[2] == "Private" else RED
    for c in range(1, 10):
        ws.cell(row=i, column=c).fill = fill

ws.cell(row=11, column=1, value="Note: SBI ties HDFC on ranks — only PSU competitive across every dimension. PNB still last.").font = ITAL
ws.merge_cells("A11:I11")
autosize(ws, min_w=12, max_w=22)

# ===================== SAVE =====================
wb.save(OUT)
print("wrote", OUT)
